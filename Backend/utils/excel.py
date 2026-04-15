"""Excel export utilities using openpyxl."""

import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _header_style(ws, row: int, headers: list[str]):
    fill = PatternFill("solid", fgColor="1E6FBA")
    font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_students(students: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách sinh viên"

    headers = ["MSSV", "Họ tên", "Ngày sinh", "Giới tính", "Lớp", "Khoa", "Email", "SĐT", "Trạng thái", "GPA"]
    _header_style(ws, 1, headers)

    for sv in students:
        ws.append([
            sv.get("mssv"), sv.get("ho_ten"),
            sv.get("ngay_sinh"), sv.get("gioi_tinh"),
            sv.get("lop"), sv.get("khoa"),
            sv.get("email"), sv.get("so_dien_thoai"),
            sv.get("trang_thai"), sv.get("gpa"),
        ])
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_grades(grades: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"

    headers = ["MSSV", "Họ tên", "Mã HP", "Tên học phần", "Tín chỉ", "Học kỳ", "ĐGK", "ĐCK", "Tổng kết", "Kết quả"]
    _header_style(ws, 1, headers)

    for g in grades:
        ws.append([
            g.get("mssv"), g.get("ho_ten"), g.get("ma_hp"), g.get("ten_hp"),
            g.get("so_tin_chi"), g.get("hoc_ky"),
            g.get("diem_gk"), g.get("diem_ck"),
            g.get("tong_ket"), g.get("ket_qua"),
        ])
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_debts(debts: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sinh viên nợ học phí"

    headers = ["MSSV", "Họ tên", "Phải nộp", "Đã nộp", "Còn thiếu", "Hạn nộp", "Trạng thái"]
    _header_style(ws, 1, headers)

    for d in debts:
        phai_nop = d.get("phai_nop", 0)
        da_nop = d.get("da_nop", 0)
        ws.append([
            d.get("mssv"), d.get("ho_ten"),
            phai_nop, da_nop,
            phai_nop - da_nop,
            d.get("han_nop"), d.get("trang_thai"),
        ])
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
